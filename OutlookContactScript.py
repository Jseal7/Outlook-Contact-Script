import tkinter.font
import win32com.client
from openpyxl import Workbook, load_workbook
import tkinter
from tkinter import *

def getOutlookCOntacts():
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    contacts = outlook.GetDefaultFolder(10).Items
    contacts_list = []

    for contact in contacts:
        contact_info = {}

        contact_info['Name'] = contact.FullName

        contact_info['Email1'] = contact.Email1Address
        contact_info['Email2'] = contact.Email2Address
        contact_info['Email3'] = contact.Email3Address

        contact_info['Business'] = contact.BusinessTelephoneNumber
        contact_info['Home'] = contact.HomeTelephoneNumber
        contact_info['Mobile'] = contact.MobileTelephoneNumber

        contact_info['Address'] = contact.MailingAddress

        contact_info['Company'] = contact.CompanyName
        contact_info['Job Title'] = contact.JobTitle
        
        contacts_list.append(contact_info)

    return contacts_list

def makeExcel(contacts_list, filename):
    excelBook = Workbook()
    currSheet = excelBook.active
    currSheet.title = "Outlook Contacts"
    currSheet.append(['Name', 'Email1', 'Email2', 'Email3', 'Business', 'Home', 'Mobile', 'Address', 'Company', 'Job Title'])
    
    for contact in contacts_list:
        currSheet.append([contact['Name'], contact['Email1'], contact['Email2'], contact['Email3'], contact['Business'], contact['Home'], contact['Mobile'], contact["Address"], contact['Company'], contact['Job Title']])
    
    excelBook.save(filename)

def makeContacts(filename, contacts_list):
    try:
        excelBook = load_workbook(filename)
        currSheet = excelBook.active
    except:
        excelBook = Workbook()
        currSheet = excelBook.active
        excelBook.save(filename)


    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    contacts = outlook.GetDefaultFolder(10).Items

    nameArr = []
    for i in range(len(contacts_list)):
        nameArr.append(contacts_list[i]['Name'])
    
    for i in range(2, currSheet.max_row + 1):
        if (not (currSheet.cell(row= i, column= 1).value in nameArr)):
            contactItem = contacts.Add("IPM.Contact")

            if (currSheet.cell(row= i, column= 1).value != None):
                contactItem.FullName = currSheet.cell(row= i, column= 1).value
            if (currSheet.cell(row= i, column= 2).value != None):
                contactItem.Email1Address = currSheet.cell(row= i, column= 2).value
            if (currSheet.cell(row= i, column= 3).value != None):
                contactItem.Email2Address = currSheet.cell(row= i, column= 3).value
            if (currSheet.cell(row= i, column= 4).value != None):
                contactItem.Email3Address = currSheet.cell(row= i, column= 4).value    
            if (currSheet.cell(row= i, column= 5).value != None):
                contactItem.BusinessTelephoneNumber = currSheet.cell(row= i, column= 5).value
            if (currSheet.cell(row= i, column= 6).value != None):
                contactItem.HomeTelephoneNumber = currSheet.cell(row= i, column= 6).value
            if (currSheet.cell(row= i, column= 7).value != None):
                contactItem.MobileTelephoneNumber = currSheet.cell(row= i, column= 7).value
            if (currSheet.cell(row= i, column= 8).value != None):
                contactItem.MailingAddress = currSheet.cell(row= i, column= 8).value
            if (currSheet.cell(row= i, column= 9).value != None):
                contactItem.CompanyName = currSheet.cell(row= i, column= 9).value
            if (currSheet.cell(row= i, column= 10).value != None):
                contactItem.JobTitle = currSheet.cell(row= i, column= 10).value

            contactItem.Save()

def makeGui():
    guiWindow = tkinter.Tk()
    guiWindow.title("Excel to Outlook Contact")
    guiWindow.configure(background='#02D7FF')

    frame = tkinter.Frame(guiWindow, padx=20, pady=30)
    frame.pack(padx=5, pady=5)
    frame.configure(background='#000000')

    label = tkinter.Label(frame, text="Contact Functions", background="#000000", fg='#fff', font=('Tahoma', 15, 'bold'))
    label.grid(row=0, column=0, ipadx=10, pady=15)

    excelToOutlook = tkinter.Button(
        frame,
        text="Make Contacts",
        background='#02D7FF',
        activebackground='#24AAFF',
        highlightthickness=2,
        highlightbackground='#02D7FF',
        highlightcolor='#FFFFFF',
        cursor='hand2',
        command=lambda: makeContacts('outlook_contacts.xlsx', getOutlookCOntacts()))
    excelToOutlook.grid(row=1, column=0, ipadx=10, pady=5)

    outlookToExcel = tkinter.Button(
        frame, text = "Make Excel",
        background='#02D7FF',
        activebackground='#24AAFF',
        highlightthickness=2,
        highlightbackground='#02D7FF',
        highlightcolor='#FFFFFF',
        cursor='hand2',
        command=lambda: makeExcel(getOutlookCOntacts(), 'outlook_contacts.xlsx'))
    outlookToExcel.grid(row=2, column=0, ipadx=10, pady=5)

    quitButton = tkinter.Button(
        frame,
        text="Quit",
        background="#FF7F7F",
        activebackground='#D50101',
        highlightthickness=2,
        highlightbackground='#02D7FF',
        highlightcolor='#FFFFFF',
        cursor='hand2',
        command=guiWindow.quit)
    quitButton.grid(row=3, column=0, ipadx=10, pady=5)

    guiWindow.mainloop()


if __name__ == "__main__":
    makeGui()